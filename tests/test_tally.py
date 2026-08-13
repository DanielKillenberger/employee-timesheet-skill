"""Final tally: confirmation gate, snapshot-only content, templates (AC9/AC10).

Three things are being defended here:

* a tally can only ever come from confirmed hours (AC9);
* everything it prints comes from the frozen snapshot, so re-registering the
  worker afterwards cannot change a document that was already produced;
* a user template is honoured or rejected with a specific reason — never
  silently swapped for the built-in one (AC10).
"""

from __future__ import annotations

import json
import subprocess
import sys
from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from pypdf import PdfReader

import timesheet
from lib import extraction as ex
from lib import registry as reg
from lib import tally as tl
from lib.datadir import resolve_data_dir
from lib.errors import TimesheetError
from lib.layout import build_month_layout

SCRIPTS_DIR = Path(timesheet.__file__).resolve().parent
REPO_ROOT = SCRIPTS_DIR.parent
MONTH = "2026-08"  # 31 days; the worker below works Mondays and Tuesdays


@pytest.fixture()
def data_dir(tmp_path: Path):
    directory = resolve_data_dir(str(tmp_path / "data"))
    reg.register_worker(
        directory,
        worker_id="anna",
        display_name="Anna Muster",
        weekdays="mon,tue",
        rate="7.50",
    )
    return directory


def entries_document(hours: str = "7.5", observed_name: str = "Anna Muster") -> dict:
    """One value entry per working day of 2026-08 (Mondays and Tuesdays)."""
    working = [3, 4, 10, 11, 17, 18, 24, 25, 31]
    return {
        "schema_version": 1,
        "observed_name": {"kind": "value", "value": observed_name},
        "entries": [
            {"date": f"2026-08-{day:02d}", "kind": "value", "value": hours, "confidence": "high"}
            for day in working
        ],
    }


def make_session(data_dir, *, confirm: bool = True, hours: str = "7.5"):
    record = reg.get_worker(data_dir, "anna")
    layout = build_month_layout(record, MONTH)
    # The sheet always carries the currently registered name, so the identity
    # gate never gets in the way of what these tests are actually checking.
    document = ex.parse_entries_document(
        entries_document(hours, observed_name=record["display_name"]),
        layout=layout,
        worker_id="anna",
    )
    session = ex.build_session(record=record, layout=layout, document=document)
    if confirm:
        ex.confirm_session(session, record)
    ex.save_session(data_dir, session)
    return session


def tally(data_dir, **kwargs):
    return tl.generate_tally(
        data_dir,
        worker_id="anna",
        month=MONTH,
        generated_on=date(2026, 9, 1),
        **kwargs,
    )


def pdf_text(path: str | Path) -> str:
    return "\n".join(page.extract_text() or "" for page in PdfReader(str(path)).pages)


def run_cli(*args: str, data_dir) -> dict:
    completed = subprocess.run(
        [sys.executable, str(SCRIPTS_DIR / "timesheet.py"), *args, "--data-dir", str(data_dir.path), "--json"],
        capture_output=True,
        text=True,
        check=False,
    )
    stream = completed.stdout if completed.returncode == 0 else completed.stderr
    return {"returncode": completed.returncode, "payload": json.loads(stream)}


# --------------------------------------------------------------------------
# AC9 — no tally without confirmation
# --------------------------------------------------------------------------


def test_an_unconfirmed_session_is_refused(data_dir) -> None:
    make_session(data_dir, confirm=False)
    with pytest.raises(TimesheetError) as error:
        tally(data_dir)
    assert error.value.code == "not_confirmed"
    assert "confirm" in error.value.message
    assert list(data_dir.output_dir.iterdir()) == []


def test_a_missing_session_is_refused_by_name(data_dir) -> None:
    with pytest.raises(TimesheetError) as error:
        tally(data_dir)
    assert error.value.code == "no_session"


def test_the_cli_refuses_an_unconfirmed_month(data_dir) -> None:
    make_session(data_dir, confirm=False)
    result = run_cli("tally", "--worker", "anna", "--month", MONTH, data_dir=data_dir)
    assert result["returncode"] != 0
    assert result["payload"]["code"] == "not_confirmed"


# --------------------------------------------------------------------------
# AC9 — the built-in PDF, always, complete, from the snapshot only
# --------------------------------------------------------------------------


def test_the_builtin_pdf_carries_every_required_fact(data_dir) -> None:
    make_session(data_dir)
    result = tally(data_dir)
    text = pdf_text(result["files"]["pdf"])

    assert "Anna Muster" in text
    assert "August 2026" in text
    assert "01.09.2026" in text  # generation date
    assert "67.5" in text  # 9 working days x 7.5 h
    assert "7.50" in text and "CHF" in text
    assert "506.25" in text  # 67.5 x 7.50
    assert "67.5 h × 7.50 CHF = 506.25 CHF" in text
    assert tl.TALLY_TITLE in text

    for day in result["days"]:
        assert day["label"] in text
    assert len(result["days"]) == 31  # every calendar date, not only working days


def test_the_builtin_pdf_is_a_single_a4_page(data_dir) -> None:
    make_session(data_dir)
    result = tally(data_dir)
    assert len(PdfReader(result["files"]["pdf"]).pages) == 1


def test_rounding_is_shown_transparently_when_it_changed_the_amount(data_dir) -> None:
    reg.register_worker(data_dir, worker_id="anna", rate="7.3333")
    make_session(data_dir, hours="7.5")
    result = tally(data_dir)

    assert result["rounding_applied"] is True
    assert result["exact_amount"] == "494.99775"
    assert result["gross_pay"] == "495.00"
    assert tl.ROUNDING_NOTE in result["calculation"]
    assert "494.99775" in pdf_text(result["files"]["pdf"])


def test_the_tally_reads_the_snapshot_and_not_the_registry(data_dir) -> None:
    """End-to-end immutability: re-registering after confirmation changes nothing."""
    make_session(data_dir)
    before = tally(data_dir)
    before_text = pdf_text(before["files"]["pdf"])

    reg.register_worker(
        data_dir, worker_id="anna", display_name="Anna Neu", rate="99.00", currency="EUR"
    )
    after = tally(data_dir, reserve=None)

    assert after["worker"]["display_name"] == "Anna Muster"
    assert after["hourly_rate"] == "7.50"
    assert after["currency"] == "CHF"
    assert after["gross_pay"] == before["gross_pay"]
    assert after["calculation"] == before["calculation"]
    assert pdf_text(after["files"]["pdf"]) == before_text


def test_existing_documents_are_not_replaced_without_force(data_dir) -> None:
    make_session(data_dir)
    from lib.datadir import reserve_new_file

    first = tally(data_dir, reserve=reserve_new_file)
    original = Path(first["files"]["pdf"]).read_bytes()

    with pytest.raises(TimesheetError) as error:
        tally(data_dir, reserve=reserve_new_file)
    assert error.value.code == "output_exists"
    assert Path(first["files"]["pdf"]).read_bytes() == original


# --------------------------------------------------------------------------
# AC10 — templates: default, stored, explicit
# --------------------------------------------------------------------------


def test_the_bundled_default_template_is_tracked_in_git() -> None:
    """A gitignored default would leave an installed skill with no template.

    ``*.xlsx`` is ignored globally, so this asserts the exception actually
    holds — from git's own file listing, not from the file existing on disk.
    """
    relative = "assets/default-tally-template.xlsx"
    assert (REPO_ROOT / relative).is_file()

    listed = subprocess.run(
        ["git", "ls-files", "--", relative],
        cwd=REPO_ROOT,
        capture_output=True,
        text=True,
        check=True,
    ).stdout.split()
    assert listed == [relative], "the bundled template is not tracked by git"

    ignored = subprocess.run(
        ["git", "check-ignore", "--no-index", "-q", "--", relative],
        cwd=REPO_ROOT,
        capture_output=True,
        text=True,
        check=False,
    )
    assert ignored.returncode == 1, "the bundled template is still matched by a .gitignore rule"


def test_the_default_template_works_out_of_the_box(data_dir) -> None:
    make_session(data_dir)
    result = tally(data_dir)

    assert result["template"]["source"] == tl.TEMPLATE_SOURCE_BUNDLED
    assert result["template"]["path"] == str(tl.BUNDLED_TEMPLATE)

    sheet = load_workbook(result["files"]["xlsx"]).active
    values = [tuple(row) for row in sheet.iter_rows(values_only=True)]
    flat = [cell for row in values for cell in row if cell is not None]

    assert "Anna Muster" in flat
    assert "August 2026" in flat
    assert "01.09.2026" in flat
    assert "CHF" in flat
    assert 67.5 in flat  # total hours, as a real number
    assert 506.25 in flat  # gross pay, as a real number
    assert not any(isinstance(cell, str) and "{{" in cell for cell in flat)


def test_every_confirmed_day_becomes_a_row_in_the_template(data_dir) -> None:
    session = make_session(data_dir)
    result = tally(data_dir)
    sheet = load_workbook(result["files"]["xlsx"]).active

    labels = {row[0] for row in sheet.iter_rows(values_only=True) if isinstance(row[0], str)}
    for day in session["confirmation"]["days"]:
        assert day["label"] in labels

    hours_by_label = {
        row[0]: row[1] for row in sheet.iter_rows(values_only=True) if isinstance(row[0], str)
    }
    assert hours_by_label["Mo, 03.08.2026"] == 7.5
    assert hours_by_label["Sa, 01.08.2026"] == 0


def test_content_below_the_marker_row_moves_down_intact(data_dir) -> None:
    """The total line under ``{{day_rows}}`` must survive the row expansion."""
    make_session(data_dir)
    result = tally(data_dir)
    sheet = load_workbook(result["files"]["xlsx"]).active
    rows = [tuple(row) for row in sheet.iter_rows(values_only=True)]

    total_row = next(index for index, row in enumerate(rows) if row[0] == "Total Stunden")
    first_day_row = next(index for index, row in enumerate(rows) if row[0] == "Mo, 03.08.2026")
    assert total_row > first_day_row
    assert rows[total_row][1] == 67.5

    # The merged disclaimer at the bottom moved with everything else.
    merged = {str(item) for item in sheet.merged_cells.ranges}
    assert "A1:B1" in merged
    assert any(item != "A1:B1" for item in merged)


def test_an_explicit_template_beats_the_stored_and_bundled_ones(data_dir, tmp_path: Path) -> None:
    make_session(data_dir)
    explicit = write_template(tmp_path / "mine.xlsx", title="MEINE VORLAGE")
    stored = data_dir.child("templates", "tally.xlsx")
    stored.parent.mkdir(parents=True, exist_ok=True)
    write_template(stored, title="GESPEICHERT")

    result = tally(data_dir, template=str(explicit))
    assert result["template"]["source"] == tl.TEMPLATE_SOURCE_ARGUMENT
    flat = [
        cell
        for row in load_workbook(result["files"]["xlsx"]).active.iter_rows(values_only=True)
        for cell in row
    ]
    assert "MEINE VORLAGE" in flat
    assert "GESPEICHERT" not in flat


def test_a_stored_template_beats_the_bundled_one(data_dir) -> None:
    make_session(data_dir)
    stored = data_dir.child("templates", "tally.xlsx")
    stored.parent.mkdir(parents=True, exist_ok=True)
    write_template(stored, title="GESPEICHERT")

    result = tally(data_dir)
    assert result["template"]["source"] == tl.TEMPLATE_SOURCE_DATA_DIR
    assert result["template"]["path"] == str(stored)


def test_the_builtin_pdf_is_still_produced_with_a_user_template(data_dir, tmp_path: Path) -> None:
    """R5: the mandatory PDF never depends on the template or on a converter."""
    make_session(data_dir)
    result = tally(data_dir, template=str(write_template(tmp_path / "mine.xlsx")))

    assert Path(result["files"]["pdf"]).is_file()
    assert len(PdfReader(result["files"]["pdf"]).pages) == 1
    assert "Anna Muster" in pdf_text(result["files"]["pdf"])


# --------------------------------------------------------------------------
# AC10 — a broken template is an error naming the problem
# --------------------------------------------------------------------------


def write_template(path: Path, *, title: str = "VORLAGE", omit: str | None = None, extra_marker: bool = False) -> Path:
    """A minimal but valid template, optionally with one placeholder removed."""
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = title
    row = 2
    for placeholder in tl.SCALAR_PLACEHOLDERS:
        if placeholder == omit:
            continue
        sheet.cell(row=row, column=1, value=placeholder.strip("{}"))
        sheet.cell(row=row, column=2, value=placeholder)
        row += 1
    if omit != tl.PLACEHOLDER_DAY_ROWS:
        sheet.cell(row=row, column=1, value=tl.PLACEHOLDER_DAY_ROWS)
        row += 1
    if extra_marker:
        sheet.cell(row=row + 3, column=1, value=tl.PLACEHOLDER_DAY_ROWS)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


@pytest.mark.parametrize(
    "placeholder",
    [
        tl.PLACEHOLDER_WORKER_NAME,
        tl.PLACEHOLDER_MONTH_TITLE,
        tl.PLACEHOLDER_GENERATION_DATE,
        tl.PLACEHOLDER_TOTAL_HOURS,
        tl.PLACEHOLDER_RATE,
        tl.PLACEHOLDER_CURRENCY,
        tl.PLACEHOLDER_GROSS_PAY,
        tl.PLACEHOLDER_DAY_ROWS,
    ],
)
def test_a_missing_placeholder_is_named_and_nothing_falls_back(
    data_dir, tmp_path: Path, placeholder: str
) -> None:
    make_session(data_dir)
    broken = write_template(tmp_path / "broken.xlsx", omit=placeholder)

    with pytest.raises(TimesheetError) as error:
        tally(data_dir, template=str(broken))

    assert error.value.code == "template_placeholder_missing"
    assert placeholder in error.value.message
    assert error.value.detail["missing"] == [placeholder]
    # No half-written tally survives the refusal.
    assert not list(data_dir.output_dir.glob("*.xlsx"))


def test_two_day_row_markers_are_rejected(data_dir, tmp_path: Path) -> None:
    make_session(data_dir)
    ambiguous = write_template(tmp_path / "twice.xlsx", extra_marker=True)
    with pytest.raises(TimesheetError) as error:
        tally(data_dir, template=str(ambiguous))
    assert error.value.code == "template_placeholder_ambiguous"
    assert tl.PLACEHOLDER_DAY_ROWS in error.value.message


def test_an_unreadable_template_is_reported_as_such(data_dir, tmp_path: Path) -> None:
    make_session(data_dir)
    not_a_workbook = tmp_path / "notes.xlsx"
    not_a_workbook.write_text("this is not a spreadsheet", encoding="utf-8")

    with pytest.raises(TimesheetError) as error:
        tally(data_dir, template=str(not_a_workbook))
    assert error.value.code == "template_unreadable"
    assert str(not_a_workbook) in error.value.message


def test_a_template_path_that_does_not_exist_is_reported(data_dir, tmp_path: Path) -> None:
    make_session(data_dir)
    with pytest.raises(TimesheetError) as error:
        tally(data_dir, template=str(tmp_path / "nope.xlsx"))
    assert error.value.code == "template_missing"


# --------------------------------------------------------------------------
# formula-injection safety (plan decision 12)
# --------------------------------------------------------------------------


@pytest.mark.parametrize("hostile", ["=HYPERLINK(\"http://evil\")", "+1+1", "-2+3", "@SUM(A1)"])
def test_adversarial_names_land_in_the_template_as_literal_text(
    data_dir, tmp_path: Path, hostile: str
) -> None:
    reg.register_worker(data_dir, worker_id="anna", display_name=hostile)
    make_session(data_dir)
    result = tally(data_dir, template=str(write_template(tmp_path / "mine.xlsx")))

    sheet = load_workbook(result["files"]["xlsx"]).active
    cell = next(
        cell
        for row in sheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value == hostile
    )
    assert cell.data_type == "s"
    assert cell.quotePrefix is True


def test_an_adversarial_currency_label_is_text_too(data_dir, tmp_path: Path) -> None:
    reg.register_worker(data_dir, worker_id="anna", currency="=1+1")
    make_session(data_dir)
    result = tally(data_dir, template=str(write_template(tmp_path / "mine.xlsx")))
    sheet = load_workbook(result["files"]["xlsx"]).active
    cell = next(
        cell for row in sheet.iter_rows() for cell in row if cell.value == "=1+1"
    )
    assert cell.data_type == "s"


# --------------------------------------------------------------------------
# the optional templated PDF
# --------------------------------------------------------------------------


def test_a_missing_converter_is_reported_and_never_blocks_the_tally(data_dir, monkeypatch) -> None:
    make_session(data_dir)
    monkeypatch.setattr(tl, "find_converter", lambda: None)
    result = tally(data_dir)

    assert result["files"]["templated_pdf"] is None
    assert result["notes"] and "LibreOffice" in result["notes"][0]
    assert Path(result["files"]["pdf"]).is_file()  # the mandatory PDF is unaffected
    assert Path(result["files"]["xlsx"]).is_file()
    assert not list(data_dir.output_dir.glob("*vorlage*"))


def test_a_failing_converter_is_reported_and_never_blocks_the_tally(data_dir, monkeypatch) -> None:
    make_session(data_dir)
    monkeypatch.setattr(tl, "find_converter", lambda: "/bin/false")
    result = tally(data_dir)

    assert result["files"]["templated_pdf"] is None
    assert result["notes"] and "LibreOffice" in result["notes"][0]
    assert Path(result["files"]["pdf"]).is_file()


def test_a_working_converter_produces_a_separate_templated_pdf(data_dir, tmp_path: Path) -> None:
    """The conversion must not land on the built-in PDF's filename."""
    make_session(data_dir)
    fake = tmp_path / "fake-soffice"
    fake.write_text(
        "#!/bin/sh\n"
        # Mimic LibreOffice: write <stem>.pdf into --outdir.
        'outdir=""\n'
        'while [ $# -gt 0 ]; do case "$1" in --outdir) outdir="$2"; shift 2;; '
        '*.xlsx) src="$1"; shift;; *) shift;; esac; done\n'
        'base=$(basename "$src" .xlsx)\n'
        'printf "%%PDF-1.4\\n%%%%EOF\\n" > "$outdir/$base.pdf"\n',
        encoding="utf-8",
    )
    fake.chmod(0o755)

    result = tally(data_dir, converter=str(fake))
    templated = result["files"]["templated_pdf"]

    assert templated is not None
    assert Path(templated).is_file()
    assert templated != result["files"]["pdf"]
    assert Path(result["files"]["pdf"]).read_bytes().startswith(b"%PDF")
    assert len(PdfReader(result["files"]["pdf"]).pages) == 1  # untouched by the converter
    assert result["notes"] == []


# --------------------------------------------------------------------------
# the CLI surface
# --------------------------------------------------------------------------


def test_the_cli_reports_every_file_it_wrote(data_dir) -> None:
    make_session(data_dir)
    result = run_cli("tally", "--worker", "anna", "--month", MONTH, data_dir=data_dir)
    assert result["returncode"] == 0

    payload = result["payload"]
    assert payload["command"] == "tally"
    assert Path(payload["files"]["pdf"]).is_file()
    assert Path(payload["files"]["xlsx"]).is_file()
    assert payload["template"]["source"] == tl.TEMPLATE_SOURCE_BUNDLED
    assert payload["total_hours"] == "67.5"
    assert payload["gross_pay"] == "506.25"
    assert "67.5 h × 7.50 CHF = 506.25 CHF" == payload["calculation"]
    # A converter note, if any, is surfaced to the user as a warning.
    assert payload["notes"] == [] or payload["notes"][0] in payload["warnings"]


def test_the_cli_refuses_to_overwrite_without_force(data_dir) -> None:
    make_session(data_dir)
    assert run_cli("tally", "--worker", "anna", "--month", MONTH, data_dir=data_dir)["returncode"] == 0

    again = run_cli("tally", "--worker", "anna", "--month", MONTH, data_dir=data_dir)
    assert again["returncode"] != 0
    assert again["payload"]["code"] == "output_exists"

    forced = run_cli("tally", "--worker", "anna", "--month", MONTH, "--force", data_dir=data_dir)
    assert forced["returncode"] == 0


def test_a_partly_claimed_output_set_leaves_no_empty_files_behind(data_dir) -> None:
    """A refusal must not leave the empty claims of the files it did reserve."""
    from lib.datadir import reserve_new_file

    make_session(data_dir)
    existing = data_dir.output_dir / tl.tally_xlsx_filename("anna", MONTH)
    existing.write_bytes(b"already here")

    with pytest.raises(TimesheetError) as error:
        tally(data_dir, reserve=reserve_new_file)
    assert error.value.code == "output_exists"
    assert [item.name for item in data_dir.output_dir.iterdir()] == [existing.name]
    assert existing.read_bytes() == b"already here"


def test_a_four_decimal_rate_is_never_displayed_rounded(data_dir) -> None:
    """A 2-decimal cell format would show 7.3333 as 7.33 — not in a pay document."""
    reg.register_worker(data_dir, worker_id="anna", rate="7.3333")
    make_session(data_dir)
    result = tally(data_dir)

    sheet = load_workbook(result["files"]["xlsx"]).active
    cells = {cell.value: cell for row in sheet.iter_rows() for cell in row}
    assert cells[7.3333].number_format == "General"
    assert cells[float(result["gross_pay"])].number_format == tl.MONEY_NUMBER_FORMAT
    assert "7.3333" in pdf_text(result["files"]["pdf"])
