"""XLSX monthly sheet: dates, grey rows, SUM formula, page setup, injection safety."""

from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

import timesheet
from lib import registry as reg
from lib.datadir import resolve_data_dir
from lib.errors import TimesheetError
from lib.layout import build_month_layout
from lib.xlsx_sheet import COLUMN_HEADERS, LABEL_RATE, LABEL_TOTAL, build_workbook, write_month_sheet

SCRIPTS_DIR = Path(timesheet.__file__).resolve().parent
GREY = "FFD9D9D9"


def worker(**overrides):
    record = {
        "id": "anna",
        "display_name": "Anna Muster",
        "working_weekdays": ["mon", "tue", "wed", "thu", "fri"],
        "hourly_rate": "7.50",
        "currency": "CHF",
        "month_overrides": {},
    }
    record.update(overrides)
    return record


def sheet(record=None, month="2027-03", **kwargs):
    layout = build_month_layout(record or worker(), month)
    return build_workbook(layout, **kwargs).active, layout


def written_sheet(tmp_path: Path, record=None, month="2027-03", **kwargs):
    """Round-trip through a real file — assertions must survive serialization."""
    layout = build_month_layout(record or worker(), month)
    path = tmp_path / "sheet.xlsx"
    write_month_sheet(layout, path, **kwargs)
    return load_workbook(path).active, layout


def day_rows(worksheet):
    """Table body rows: everything between the header row and the total row."""
    header_row = next(
        cell.row
        for cell in worksheet["A"]
        if cell.value == COLUMN_HEADERS[0]
    )
    total_row = next(cell.row for cell in worksheet["A"] if cell.value == LABEL_TOTAL)
    return list(range(header_row + 1, total_row)), header_row, total_row


# --------------------------------------------------------------------------
# AC2/AC6 — dates
# --------------------------------------------------------------------------


@pytest.mark.parametrize(("month", "expected_days"), [("2027-03", 31), ("2028-02", 29), ("2027-02", 28), ("2027-04", 30)])
def test_sheet_lists_every_date_exactly_once(tmp_path: Path, month: str, expected_days: int) -> None:
    worksheet, layout = written_sheet(tmp_path, month=month)
    body, _, _ = day_rows(worksheet)
    assert len(body) == expected_days
    labels = [worksheet.cell(row=row, column=1).value for row in body]
    assert labels == [row.label for row in layout.rows]
    assert len(set(labels)) == expected_days


def test_hour_cells_are_blank_on_working_days(tmp_path: Path) -> None:
    worksheet, layout = written_sheet(tmp_path)
    body, _, _ = day_rows(worksheet)
    for row_number, day in zip(body, layout.rows):
        assert worksheet.cell(row=row_number, column=2).value is None, day.iso


# --------------------------------------------------------------------------
# AC3 — grey rows match resolved off days
# --------------------------------------------------------------------------


def test_grey_rows_are_exactly_the_resolved_off_days(tmp_path: Path) -> None:
    record = worker(month_overrides={"2027-03": {"off": ["2027-03-02"], "extra": ["2027-03-06"]}})
    worksheet, layout = written_sheet(tmp_path, record)
    body, _, _ = day_rows(worksheet)

    greyed = set()
    for row_number, day in zip(body, layout.rows):
        fills = {worksheet.cell(row=row_number, column=column).fill.fgColor.rgb for column in (1, 2, 3)}
        if fills == {GREY}:
            greyed.add(day.iso)
        else:
            assert GREY not in fills, f"row {day.iso} is only partly grey"

    assert greyed == {row.iso for row in layout.off_rows}
    assert "2027-03-02" in greyed  # override off beats the weekday schedule
    assert "2027-03-06" not in greyed  # override extra beats the weekend


def test_a_single_fill_object_is_reused_for_every_grey_cell() -> None:
    worksheet, _ = sheet()
    fills = {
        id(worksheet.cell(row=row, column=1).fill)
        for row in range(1, worksheet.max_row + 1)
        if worksheet.cell(row=row, column=1).fill.fgColor.rgb == GREY
    }
    assert len(fills) == 1


# --------------------------------------------------------------------------
# total row — a real formula, verified by reading the file back
# --------------------------------------------------------------------------


def test_total_row_holds_a_real_sum_formula_over_the_day_rows(tmp_path: Path) -> None:
    worksheet, _ = written_sheet(tmp_path)
    body, _, total_row = day_rows(worksheet)
    cell = worksheet.cell(row=total_row, column=2)
    assert cell.data_type == "f"
    assert cell.value == f"=SUM(B{body[0]}:B{body[-1]})"


def test_sum_formula_covers_the_whole_month_in_a_leap_february(tmp_path: Path) -> None:
    worksheet, layout = written_sheet(tmp_path, month="2028-02")
    body, _, total_row = day_rows(worksheet)
    assert len(body) == 29 == len(layout.rows)
    assert worksheet.cell(row=total_row, column=2).value == f"=SUM(B{body[0]}:B{body[-1]})"


# --------------------------------------------------------------------------
# header block and the rate flag (R2)
# --------------------------------------------------------------------------


def test_rate_is_omitted_unless_explicitly_requested(tmp_path: Path) -> None:
    worksheet, _ = written_sheet(tmp_path, hourly_rate="7.50", currency="CHF")
    values = [cell.value for column in worksheet.iter_cols() for cell in column]
    assert LABEL_RATE not in values
    assert "7.50 CHF" not in values


def test_rate_is_printed_when_requested(tmp_path: Path) -> None:
    worksheet, _ = written_sheet(tmp_path, include_rate=True, hourly_rate="7.50", currency="CHF")
    values = [cell.value for column in worksheet.iter_cols() for cell in column]
    assert LABEL_RATE in values
    assert "7.50 CHF" in values


def test_header_shows_worker_name_and_german_month(tmp_path: Path) -> None:
    worksheet, _ = written_sheet(tmp_path)
    values = [cell.value for column in worksheet.iter_cols() for cell in column]
    assert "Anna Muster" in values
    assert "März 2027" in values


# --------------------------------------------------------------------------
# page setup — one portrait A4 page
# --------------------------------------------------------------------------


def test_page_setup_pins_one_portrait_a4_page(tmp_path: Path) -> None:
    worksheet, _ = written_sheet(tmp_path)
    setup = worksheet.page_setup
    assert setup.orientation == "portrait"
    assert str(setup.paperSize) == "9"  # A4
    assert setup.fitToWidth == 1
    assert setup.fitToHeight == 1
    # Excel ignores fit-to-page unless the sheet property is set as well.
    assert worksheet.sheet_properties.pageSetUpPr.fitToPage is True


def test_print_area_covers_the_whole_table(tmp_path: Path) -> None:
    worksheet, _ = written_sheet(tmp_path)
    _, _, total_row = day_rows(worksheet)
    expected = f"'{worksheet.title}'!$A$1:$C${total_row}"
    area = worksheet.print_area
    # openpyxl hands back a list when set in memory and a string when read back.
    assert (area == [expected]) or (area == expected)


def test_no_manual_page_breaks_split_the_month(tmp_path: Path) -> None:
    worksheet, _ = written_sheet(tmp_path, month="2027-12")
    assert list(worksheet.row_breaks.brk) == []
    assert list(worksheet.col_breaks.brk) == []


# --------------------------------------------------------------------------
# plan decision 12 — formula-injection safety
# --------------------------------------------------------------------------


@pytest.mark.parametrize("hostile", ["=SUM(1,1)", "+1+1", "-1-1", "@SUM(A1)", '=HYPERLINK("http://x","x")'])
def test_user_derived_strings_are_written_as_literal_text(tmp_path: Path, hostile: str) -> None:
    record = worker(display_name=hostile, currency=hostile)
    worksheet, _ = written_sheet(
        tmp_path, record, include_rate=True, hourly_rate="7.50", currency=hostile
    )
    hits = [
        cell
        for column in worksheet.iter_cols()
        for cell in column
        if isinstance(cell.value, str) and hostile in cell.value
    ]
    assert hits, "the hostile value should still be present, just inert"
    for cell in hits:
        assert cell.data_type == "s", f"{cell.coordinate} became a formula cell"


def test_the_only_formula_in_the_sheet_is_the_generated_total(tmp_path: Path) -> None:
    record = worker(display_name="=cmd|'/c calc'!A1")
    worksheet, _ = written_sheet(tmp_path, record)
    formulas = [
        cell.coordinate
        for column in worksheet.iter_cols()
        for cell in column
        if cell.data_type == "f"
    ]
    _, _, total_row = day_rows(worksheet)
    assert formulas == [f"B{total_row}"]


# --------------------------------------------------------------------------
# CLI — generate
# --------------------------------------------------------------------------


@pytest.fixture()
def data_dir(tmp_path: Path):
    return resolve_data_dir(str(tmp_path / "data"))


def run_cli(args: list[str], data_dir) -> tuple[int, dict, str]:
    process = subprocess.run(
        [sys.executable, str(SCRIPTS_DIR / "timesheet.py"), *args, "--data-dir", str(data_dir.path), "--json"],
        capture_output=True,
        text=True,
    )
    stream = process.stdout if process.returncode == 0 else process.stderr
    return process.returncode, json.loads(stream), process.stderr


def register_anna(data_dir, **kwargs):
    return reg.register_worker(
        data_dir,
        worker_id="anna",
        display_name="Anna Muster",
        weekdays="mon,tue,wed,thu,fri",
        rate="7,50",
        **kwargs,
    )


def test_generate_writes_the_sheet_into_the_output_folder(data_dir) -> None:
    register_anna(data_dir)
    code, payload, _ = run_cli(["generate", "--worker", "anna", "--month", "2027-03"], data_dir)
    assert code == 0
    path = Path(payload["files"]["xlsx"])
    assert path == data_dir.path / "output" / "anna-2027-03.xlsx"
    assert path.exists() and path.stat().st_size > 0
    assert (payload["days"], payload["working_days"], payload["off_days"]) == (31, 23, 8)
    assert payload["rate_printed"] is False


def test_generate_refuses_to_overwrite_without_force(data_dir) -> None:
    register_anna(data_dir)
    run_cli(["generate", "--worker", "anna", "--month", "2027-03"], data_dir)
    code, payload, _ = run_cli(["generate", "--worker", "anna", "--month", "2027-03"], data_dir)
    assert code != 0
    assert payload["code"] == "output_exists"

    code, payload, _ = run_cli(
        ["generate", "--worker", "anna", "--month", "2027-03", "--force"], data_dir
    )
    assert code == 0


def test_generate_reports_unknown_worker_and_bad_month(data_dir) -> None:
    register_anna(data_dir)
    code, payload, _ = run_cli(["generate", "--worker", "nobody", "--month", "2027-03"], data_dir)
    assert code != 0 and payload["code"] == "unknown_worker"

    code, payload, _ = run_cli(["generate", "--worker", "anna", "--month", "2027-13"], data_dir)
    assert code != 0 and payload["code"] == "invalid_month"


def test_generate_warns_when_the_month_has_no_working_day(data_dir) -> None:
    reg.register_worker(
        data_dir, worker_id="bob", display_name="Bob", weekdays="", rate="10"
    )
    code, payload, _ = run_cli(["generate", "--worker", "bob", "--month", "2027-03"], data_dir)
    assert code == 0
    assert payload["working_days"] == 0
    assert any("no working days" in warning for warning in payload["warnings"])


def test_generated_file_is_owner_only_even_with_force(data_dir) -> None:
    register_anna(data_dir)
    code, payload, _ = run_cli(
        ["generate", "--worker", "anna", "--month", "2027-03", "--force"], data_dir
    )
    assert code == 0
    mode = Path(payload["files"]["xlsx"]).stat().st_mode & 0o777
    assert mode == 0o600


def test_a_failed_regeneration_leaves_the_existing_sheet_intact(data_dir, monkeypatch) -> None:
    register_anna(data_dir)
    layout = build_month_layout(reg.get_worker(data_dir, "anna"), "2027-03")
    target = data_dir.output_dir / "anna-2027-03.xlsx"
    write_month_sheet(layout, target)
    original = target.read_bytes()

    class Boom(RuntimeError):
        pass

    def explode(self, filename):  # noqa: ANN001
        Path(filename).write_bytes(b"half-written")
        raise Boom("disk full")

    monkeypatch.setattr("openpyxl.workbook.workbook.Workbook.save", explode)
    with pytest.raises(Boom):
        write_month_sheet(layout, target)

    assert target.read_bytes() == original
    leftovers = [path.name for path in target.parent.iterdir() if path.name.startswith(".tmp-")]
    assert leftovers == []


def test_control_characters_are_refused_at_registration(data_dir) -> None:
    for field, code in (("display_name", "invalid_name"), ("currency", "invalid_currency")):
        with pytest.raises(TimesheetError) as excinfo:
            reg.register_worker(
                data_dir,
                worker_id="ctrl",
                display_name="Anna\x07Muster" if field == "display_name" else "Anna",
                weekdays="mon",
                rate="7.50",
                currency="CH\x01F" if field == "currency" else None,
            )
        assert excinfo.value.code == code


def test_a_control_character_that_reached_storage_yields_a_plain_error(tmp_path: Path) -> None:
    record = worker(display_name="Anna\x07Muster")
    layout = build_month_layout(record, "2027-03")
    with pytest.raises(TimesheetError) as excinfo:
        write_month_sheet(layout, tmp_path / "sheet.xlsx")
    assert excinfo.value.code == "unwritable_text"


def test_year_zero_is_refused_by_the_cli_with_a_structured_error(data_dir) -> None:
    register_anna(data_dir)
    code, payload, stderr = run_cli(["generate", "--worker", "anna", "--month", "0000-01"], data_dir)
    assert code != 0
    assert payload["code"] == "invalid_month"
    assert "Traceback" not in stderr


def test_generate_include_rate_flag_reaches_the_sheet(data_dir) -> None:
    register_anna(data_dir)
    code, payload, _ = run_cli(
        ["generate", "--worker", "anna", "--month", "2027-03", "--include-rate"], data_dir
    )
    assert code == 0 and payload["rate_printed"] is True
    worksheet = load_workbook(payload["files"]["xlsx"]).active
    values = [cell.value for column in worksheet.iter_cols() for cell in column]
    assert "7.50 CHF" in values
