"""Final tally for the employee: built-in PDF plus optional templated XLSX (R5/R6).

The tally is the document that says "you worked these hours and this is the
gross amount". Three rules shape this module, and all three are about not
letting a payroll number drift:

* **Confirmed input only** (AC9). :func:`load_confirmed_session` refuses any
  session that is not ``confirmed`` with a plain-language error. There is no
  flag to override it.
* **Snapshot only** (plan decision 1). Everything printed — name, rate,
  currency, per-day hours, total, gross pay — comes from the frozen
  ``confirmation`` block written at confirmation time. The registry is never
  read here, so re-registering a worker afterwards cannot retroactively change
  a tally that has already been produced.
* **The built-in PDF is unconditional** (R5). It is rendered with reportlab and
  therefore depends on nothing outside this repository. A user template
  additionally produces a filled XLSX, and a templated PDF only where
  LibreOffice happens to exist; the result reports exactly which files were
  written and why one is missing. "Honouring the template" never means quietly
  substituting the built-in document for it.

Template contract (documented for users in ``references/templates.md``):

``{{worker_name}}``, ``{{month_title}}``, ``{{generation_date}}``,
``{{total_hours}}``, ``{{rate}}``, ``{{currency}}`` and ``{{gross_pay}}`` are
replaced wherever they appear; ``{{day_rows}}`` marks a row that is cloned once
per confirmed day, taking the date label in the marker's own column and the
hours in the column to its right. Every required placeholder must be present —
a template missing one is an error naming it, never a silent fallback.
"""

from __future__ import annotations

import copy
import shutil
import subprocess
import tempfile
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, Spacer, Table

from .datadir import DataDir, atomic_output_file, atomic_output_files, safe_child
from .errors import TimesheetError
from .extraction import STATUS_CONFIRMED, load_session, month_dates
from .layout import validate_month
from .money import canonical_decimal_string, parse_hours
from .pay import ZERO, build_receipt
from .pdf_sheet import (
    NOTE_STYLE,
    TABLE_HEADER_HEIGHT,
    TABLE_ROW_HEIGHT,
    TITLE_STYLE,
    build_document,
    ensure_printable,
    escape,
    labelled_line,
    table_style,
)
from .xlsx_sheet import write_text_cell

# --------------------------------------------------------------------------
# German document wording
# --------------------------------------------------------------------------

TALLY_TITLE = "Stundenabrechnung"
LABEL_WORKER = "Mitarbeiter/in"
LABEL_MONTH = "Monat"
LABEL_GENERATED = "Erstellt am"
LABEL_TOTAL_HOURS = "Total Stunden"
LABEL_RATE = "Stundenlohn"
LABEL_GROSS = "Bruttolohn"
LABEL_CALCULATION = "Berechnung"
TALLY_COLUMN_HEADERS = ("Datum", "Stunden")
ROUNDING_NOTE = "kaufmännisch auf 0.01 gerundet"
DISCLAIMER = (
    "Bruttolohn = bestätigte Stunden × registrierter Stundenlohn. "
    "Abzüge, Zuschläge, Sozialversicherungen und Steuern sind darin nicht berücksichtigt."
)

#: Same date/hours column widths as the monthly sheet, so the two documents
#: look like a matching pair.
TALLY_COLUMN_WIDTHS = (46 * mm, 26 * mm)

# --------------------------------------------------------------------------
# template contract
# --------------------------------------------------------------------------

PLACEHOLDER_WORKER_NAME = "{{worker_name}}"
PLACEHOLDER_MONTH_TITLE = "{{month_title}}"
PLACEHOLDER_GENERATION_DATE = "{{generation_date}}"
PLACEHOLDER_TOTAL_HOURS = "{{total_hours}}"
PLACEHOLDER_RATE = "{{rate}}"
PLACEHOLDER_CURRENCY = "{{currency}}"
PLACEHOLDER_GROSS_PAY = "{{gross_pay}}"
PLACEHOLDER_DAY_ROWS = "{{day_rows}}"

#: Scalar placeholders, in the order a user most likely wants to read them.
SCALAR_PLACEHOLDERS: tuple[str, ...] = (
    PLACEHOLDER_WORKER_NAME,
    PLACEHOLDER_MONTH_TITLE,
    PLACEHOLDER_GENERATION_DATE,
    PLACEHOLDER_TOTAL_HOURS,
    PLACEHOLDER_RATE,
    PLACEHOLDER_CURRENCY,
    PLACEHOLDER_GROSS_PAY,
)
REQUIRED_PLACEHOLDERS: tuple[str, ...] = SCALAR_PLACEHOLDERS + (PLACEHOLDER_DAY_ROWS,)

#: Placeholders that become real numbers when they are a cell's entire content,
#: so a user's own template can keep calculating with them.
NUMERIC_PLACEHOLDERS: frozenset[str] = frozenset(
    {PLACEHOLDER_TOTAL_HOURS, PLACEHOLDER_RATE, PLACEHOLDER_GROSS_PAY}
)
#: Formatted with two decimals so ``1222.80`` does not read as ``1222.8``.
#: Only the gross pay qualifies: it is quantized to 0.01 by construction, so the
#: format can never hide a digit. The hourly rate keeps up to four decimals and
#: a 2-decimal format would *display* ``7.3333`` as ``7.33`` — a payroll
#: document must not round in the presentation layer.
MONEY_PLACEHOLDERS: frozenset[str] = frozenset({PLACEHOLDER_GROSS_PAY})
MONEY_NUMBER_FORMAT = "0.00"

TEMPLATE_SOURCE_ARGUMENT = "argument"
TEMPLATE_SOURCE_DATA_DIR = "data_dir"
TEMPLATE_SOURCE_BUNDLED = "bundled_default"

#: Bundled fallback template, shipped with the skill (spec R6).
BUNDLED_TEMPLATE = Path(__file__).resolve().parent.parent.parent / "assets" / "default-tally-template.xlsx"

SOFFICE_CANDIDATES: tuple[str, ...] = ("soffice", "libreoffice")
SOFFICE_TIMEOUT_SECONDS = 180


# --------------------------------------------------------------------------
# session access
# --------------------------------------------------------------------------


def _corrupt_confirmation(session: Mapping[str, Any], problem: str) -> TimesheetError:
    return TimesheetError(
        "session_corrupt",
        f"The confirmed hours for {session.get('month_title', session.get('month'))} cannot be used "
        f"because {problem}. Nothing was written. Run 'validate-extraction' with --overwrite and "
        "confirm the month again.",
        {"worker_id": session.get("worker_id"), "month": session.get("month"), "problem": problem},
    )


def validate_confirmation(session: Mapping[str, Any]) -> None:
    """Re-derive the confirmed result and refuse it if it does not add up.

    ``load_session`` checks that a confirmation has the right *shape*; that is
    not enough for the one file in this skill that becomes money. A session
    file is plain JSON in the user's own folder, so it can be edited — by a
    person, a sync conflict, a half-finished write from another tool — and
    every field the tally prints would otherwise be taken on faith.

    So everything is recomputed from the frozen per-day hours: the day set must
    be exactly the month's calendar, the total must be the exact sum, the
    receipt must be the receipt those numbers produce, and the snapshot must
    belong to the worker whose session this is. Anything else is a refusal
    naming the problem, never a document.
    """
    confirmation = session["confirmation"]
    snapshot = confirmation["snapshot"]
    month = session["month"]

    if snapshot["worker_id"] != session["worker_id"]:
        # Load-bearing: the worker id also names the output files.
        raise _corrupt_confirmation(
            session,
            f"it was frozen for worker '{snapshot['worker_id']}' but stored for "
            f"'{session['worker_id']}'",
        )

    days = confirmation["days"]
    expected_dates = month_dates(month)
    dates = [day.get("date") if isinstance(day, Mapping) else None for day in days]
    if dates != expected_dates:
        raise _corrupt_confirmation(
            session, f"it does not hold exactly the {len(expected_dates)} days of {month}"
        )

    total = ZERO
    for day in days:
        if not isinstance(day.get("label"), str) or not isinstance(day.get("working"), bool):
            raise _corrupt_confirmation(session, f"the entry for {day.get('date')} is incomplete")
        try:
            total += parse_hours(day.get("hours"))
        except TimesheetError as exc:
            raise _corrupt_confirmation(
                session, f"the hours for {day.get('date')} are not a valid number"
            ) from exc

    if canonical_decimal_string(total) != confirmation["total_hours"]:
        raise _corrupt_confirmation(
            session,
            f"its total of {confirmation['total_hours']} h does not match the "
            f"{canonical_decimal_string(total)} h in the daily entries",
        )

    try:
        expected_receipt = build_receipt(total, snapshot["hourly_rate"], snapshot["currency"])
    except TimesheetError as exc:
        raise _corrupt_confirmation(session, "its stored hourly rate is not a valid amount") from exc
    if confirmation["receipt"] != expected_receipt:
        raise _corrupt_confirmation(
            session, "the stored pay calculation does not match the confirmed hours and rate"
        )


def load_confirmed_session(data_dir: DataDir, worker_id: str, month: str) -> dict[str, Any]:
    """Load the session for ``worker_id``/``month``, refusing unconfirmed hours.

    This is the AC9 gate. A provisional extraction is exactly the situation the
    spec forbids turning into a payroll document, so the refusal names the
    command that would make it valid instead of hinting at a workaround.
    """
    session = load_session(data_dir, worker_id, validate_month(month))
    if session["status"] != STATUS_CONFIRMED:
        raise TimesheetError(
            "not_confirmed",
            f"The hours for {session.get('month_title', session['month'])} have not been confirmed yet, "
            "so no final tally can be produced. Go through the flagged days with 'confirm' first.",
            {
                "worker_id": session["worker_id"],
                "month": session["month"],
                "status": session["status"],
            },
        )
    validate_confirmation(session)
    return session


def german_date(value: date | datetime) -> str:
    """``13.08.2026`` — the date form the rest of the document uses."""
    return value.strftime("%d.%m.%Y")


def tally_context(session: Mapping[str, Any], *, generated_on: date | None = None) -> dict[str, Any]:
    """Everything the tally prints, taken exclusively from the frozen snapshot."""
    confirmation = session["confirmation"]
    snapshot = confirmation["snapshot"]
    receipt = confirmation["receipt"]
    generated = generated_on or date.today()

    return {
        # The session's own id, which resolved the session path and therefore
        # already passed the data-folder containment check. The snapshot's copy
        # is asserted equal to it in validate_confirmation.
        "worker_id": session["worker_id"],
        "worker_name": snapshot["display_name"],
        "month": session["month"],
        "month_title": session.get("month_title", session["month"]),
        "generation_date": german_date(generated),
        "total_hours": confirmation["total_hours"],
        "rate": snapshot["hourly_rate"],
        "currency": snapshot["currency"],
        "gross_pay": receipt["gross_pay"],
        "exact_amount": receipt["exact_amount"],
        "rounding_applied": bool(receipt["rounding_applied"]),
        "confirmed_at": confirmation["confirmed_at"],
        "days": [dict(day) for day in confirmation["days"]],
    }


def calculation_sentence(context: Mapping[str, Any]) -> str:
    """The transparent ``hours × rate = gross`` line, in German (R5).

    When the terminal rounding changed the number, the unrounded product is
    printed alongside it — the whole point of the receipt is that the employee
    can redo the arithmetic by hand and arrive at the same figure.
    """
    currency = context["currency"]
    sentence = (
        f"{context['total_hours']} h × {context['rate']} {currency} = {context['gross_pay']} {currency}"
    )
    if context["rounding_applied"]:
        sentence += f" (exakt {context['exact_amount']} {currency}, {ROUNDING_NOTE})"
    return sentence


# --------------------------------------------------------------------------
# built-in PDF (always produced)
# --------------------------------------------------------------------------


def build_tally_story(context: Mapping[str, Any]) -> list[Any]:
    """The flowables of the built-in tally PDF, in print order."""
    story: list[Any] = [
        Paragraph(escape(TALLY_TITLE), TITLE_STYLE),
        Spacer(1, 6),
        labelled_line(LABEL_WORKER, context["worker_name"]),
        labelled_line(LABEL_MONTH, context["month_title"]),
        labelled_line(LABEL_GENERATED, context["generation_date"]),
        Spacer(1, 8),
    ]

    rows: list[list[str]] = [list(TALLY_COLUMN_HEADERS)]
    grey_rows: list[int] = []
    for index, day in enumerate(context["days"], start=1):
        rows.append([day["label"], day["hours"]])
        if not day["working"]:
            grey_rows.append(index)
    rows.append([LABEL_TOTAL_HOURS, context["total_hours"]])
    total_row = len(rows) - 1

    heights = [TABLE_HEADER_HEIGHT] + [TABLE_ROW_HEIGHT] * (len(rows) - 1)
    table = Table(
        rows,
        colWidths=list(TALLY_COLUMN_WIDTHS),
        rowHeights=heights,
        repeatRows=0,
        hAlign="LEFT",  # flush with the heading block, like the monthly sheet
    )
    table.setStyle(table_style(grey_rows=grey_rows, total_row=total_row))
    story.append(table)

    currency = context["currency"]
    story.extend(
        [
            Spacer(1, 10),
            labelled_line(LABEL_TOTAL_HOURS, f"{context['total_hours']} h"),
            labelled_line(LABEL_RATE, f"{context['rate']} {currency}"),
            labelled_line(LABEL_GROSS, f"{context['gross_pay']} {currency}"),
            labelled_line(LABEL_CALCULATION, calculation_sentence(context)),
            Spacer(1, 8),
            Paragraph(escape(DISCLAIMER), NOTE_STYLE),
        ]
    )
    return story


def tally_pdf_filename(worker_id: str, month: str) -> str:
    return f"{worker_id}-{month}-abrechnung.pdf"


def tally_xlsx_filename(worker_id: str, month: str) -> str:
    return f"{worker_id}-{month}-abrechnung.xlsx"


def templated_pdf_filename(worker_id: str, month: str) -> str:
    # Deliberately a different stem from the built-in PDF: the two documents
    # must never be able to overwrite one another.
    return f"{worker_id}-{month}-abrechnung-vorlage.pdf"


def render_tally_pdf(context: Mapping[str, Any], path: Path) -> Path:
    """Render the mandatory built-in tally PDF (R5) straight to ``path``."""
    story = build_tally_story(context)
    document = build_document(
        path,
        title=(
            f"{TALLY_TITLE} {context['month_title']} - "
            f"{ensure_printable(context['worker_name'], subject='The name')}"
        ),
    )
    document.build(story)
    return path


def write_tally_pdf(context: Mapping[str, Any], path: Path) -> Path:
    """Write the built-in tally PDF to ``path``, replacing it atomically."""
    with atomic_output_file(path, suffix=".pdf") as tmp_path:
        render_tally_pdf(context, tmp_path)
    return path


# --------------------------------------------------------------------------
# template resolution and filling
# --------------------------------------------------------------------------


def resolve_template(data_dir: DataDir, explicit: str | None = None) -> tuple[Path, str]:
    """Resolve the tally template: argument > data dir > bundled default (R6)."""
    if explicit:
        path = Path(explicit).expanduser()
        if not path.is_file():
            raise TimesheetError(
                "template_missing",
                f"The template file '{path}' does not exist. Check the path, or leave --template "
                "away to use your stored template or the built-in default.",
                {"path": str(path)},
            )
        return path.resolve(), TEMPLATE_SOURCE_ARGUMENT

    stored = data_dir.child("templates", "tally.xlsx")
    if stored.is_file():
        return stored, TEMPLATE_SOURCE_DATA_DIR

    if not BUNDLED_TEMPLATE.is_file():
        raise TimesheetError(
            "template_missing",
            f"The built-in tally template '{BUNDLED_TEMPLATE.name}' is missing from the skill. "
            "Re-install the skill, or pass your own template with --template.",
            {"path": str(BUNDLED_TEMPLATE)},
        )
    return BUNDLED_TEMPLATE, TEMPLATE_SOURCE_BUNDLED


def _load_template(path: Path) -> Workbook:
    try:
        return load_workbook(path)
    except TimesheetError:
        raise
    except Exception as exc:  # openpyxl raises a wide family for broken files
        raise TimesheetError(
            "template_unreadable",
            f"The template '{path}' could not be read as an Excel file ({type(exc).__name__}: {exc}). "
            "Please open it in your spreadsheet program and save it again as .xlsx.",
            {"path": str(path), "reason": str(exc)},
        ) from exc


def _cell_texts(worksheet: Worksheet) -> Iterable[tuple[Any, str]]:
    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                yield cell, cell.value


def find_placeholders(workbook: Workbook) -> dict[str, list[str]]:
    """Map every known placeholder to the cell coordinates that contain it."""
    found: dict[str, list[str]] = {name: [] for name in REQUIRED_PLACEHOLDERS}
    for worksheet in workbook.worksheets:
        for cell, text in _cell_texts(worksheet):
            for name in REQUIRED_PLACEHOLDERS:
                if name in text:
                    found[name].append(f"{worksheet.title}!{cell.coordinate}")
    return found


def _require_placeholders(workbook: Workbook, template_path: Path) -> dict[str, list[str]]:
    found = find_placeholders(workbook)
    missing = [name for name in REQUIRED_PLACEHOLDERS if not found[name]]
    if missing:
        listed = ", ".join(missing)
        raise TimesheetError(
            "template_placeholder_missing",
            f"The template '{template_path}' is missing the placeholder(s) {listed}. "
            "Add a cell containing each of them and try again — the tally is not written "
            "from a template that would silently lose information.",
            {"path": str(template_path), "missing": missing, "found": {k: v for k, v in found.items() if v}},
        )
    day_row_cells = found[PLACEHOLDER_DAY_ROWS]
    if len(day_row_cells) > 1:
        raise TimesheetError(
            "template_placeholder_ambiguous",
            f"The template '{template_path}' contains {PLACEHOLDER_DAY_ROWS} more than once "
            f"({', '.join(day_row_cells)}). Keep exactly one marker row for the daily hours.",
            {"path": str(template_path), "cells": day_row_cells},
        )
    return found


def _marker_cell(workbook: Workbook) -> tuple[Worksheet, Any]:
    for worksheet in workbook.worksheets:
        for cell, text in _cell_texts(worksheet):
            if PLACEHOLDER_DAY_ROWS in text:
                return worksheet, cell
    raise AssertionError("marker presence is checked before this call")  # pragma: no cover


def _unmerge_below(worksheet: Worksheet, first_row: int) -> list[tuple[int, int, int, int]]:
    """Undo merges that start at or below ``first_row``, returning their extents.

    ``insert_rows`` moves the cell objects (styles included) but leaves merged
    ranges pointing at the coordinates they had before, and unmerging *after*
    the insert then fails because those coordinates hold different cells. So
    the merges come apart first and are put back at their new rows afterwards.
    """
    moving = [
        (merged.min_row, merged.min_col, merged.max_row, merged.max_col)
        for merged in list(worksheet.merged_cells.ranges)
        if merged.min_row >= first_row
    ]
    for min_row, min_col, max_row, max_col in moving:
        worksheet.unmerge_cells(
            start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col
        )
    return moving


def _shift_geometry_below(
    worksheet: Worksheet,
    first_row: int,
    offset: int,
    merges: list[tuple[int, int, int, int]],
) -> None:
    """Move row heights down by ``offset`` and restore ``merges`` at their new rows.

    Row heights survive ``insert_rows`` untouched, which would otherwise smear
    the template's formatting across the inserted day rows.
    """
    if offset <= 0:
        return

    heights = {
        index: dimension.height
        for index, dimension in worksheet.row_dimensions.items()
        if index >= first_row and dimension.height is not None
    }
    for index in heights:
        worksheet.row_dimensions[index].height = None
    for index, height in heights.items():
        worksheet.row_dimensions[index + offset].height = height

    for min_row, min_col, max_row, max_col in merges:
        worksheet.merge_cells(
            start_row=min_row + offset,
            start_column=min_col,
            end_row=max_row + offset,
            end_column=max_col,
        )


def _shift_print_area(worksheet: Worksheet, marker_row: int, offset: int) -> None:
    """Grow the print area to cover the rows the day table expanded into.

    A template that pins its print area to the original block would otherwise
    print only the first day: the rows move down, the print area does not, and
    the LibreOffice PDF quietly ends mid-table. A boundary *below* the marker
    row moves with the content it referred to; an end boundary *at* the marker
    row grows, because that one row became many.
    """
    if offset <= 0:
        return
    area = worksheet.print_area
    if not area:
        return
    ranges = area if isinstance(area, list) else [area]

    shifted: list[str] = []
    for entry in ranges:
        sheet_prefix, _, cells = str(entry).rpartition("!")
        try:
            min_col, min_row, max_col, max_row = range_boundaries(cells.replace("$", ""))
        except (ValueError, TypeError):  # pragma: no cover - unparseable, leave as is
            shifted.append(str(entry))
            continue
        if min_row is None or max_row is None:  # a whole-column range needs no shift
            shifted.append(str(entry))
            continue
        if min_row > marker_row:
            min_row += offset
        if max_row >= marker_row:
            max_row += offset
        moved = (
            f"${get_column_letter(min_col)}${min_row}:"
            f"${get_column_letter(max_col)}${max_row}"
        )
        shifted.append(f"{sheet_prefix}!{moved}" if sheet_prefix else moved)

    worksheet.print_area = shifted


def _expand_day_rows(workbook: Workbook, days: list[Mapping[str, Any]]) -> None:
    """Clone the ``{{day_rows}}`` marker row once per confirmed day."""
    worksheet, marker = _marker_cell(workbook)
    marker_row, date_column = marker.row, marker.column
    hours_column = date_column + 1
    extra = max(len(days) - 1, 0)

    if extra:
        # ``insert_rows`` carries the cells below the marker (and their styles)
        # down with it; the row heights and merges need moving by hand.
        merges = _unmerge_below(worksheet, marker_row + 1)
        worksheet.insert_rows(marker_row + 1, extra)
        _shift_geometry_below(worksheet, marker_row + 1, extra, merges)
        _shift_print_area(worksheet, marker_row, extra)
        for offset in range(1, extra + 1):
            for cell in worksheet[marker_row]:
                target = worksheet.cell(row=marker_row + offset, column=cell.column)
                target._style = copy.copy(cell._style)

    if not days:
        # A month with no confirmed days cannot happen (every calendar date is
        # frozen), but an empty marker row must not keep the raw placeholder.
        write_text_cell(worksheet.cell(row=marker_row, column=date_column), "")
        return

    for offset, day in enumerate(days):
        row = marker_row + offset
        write_text_cell(worksheet.cell(row=row, column=date_column), day["label"])
        # Re-parsed through the money grammar rather than trusted: the hours land
        # in the workbook as an exact Decimal, so the user's template can sum them.
        worksheet.cell(row=row, column=hours_column).value = parse_hours(day["hours"])


def _substitute_scalars(workbook: Workbook, context: Mapping[str, Any]) -> None:
    values = {
        PLACEHOLDER_WORKER_NAME: str(context["worker_name"]),
        PLACEHOLDER_MONTH_TITLE: str(context["month_title"]),
        PLACEHOLDER_GENERATION_DATE: str(context["generation_date"]),
        PLACEHOLDER_TOTAL_HOURS: str(context["total_hours"]),
        PLACEHOLDER_RATE: str(context["rate"]),
        PLACEHOLDER_CURRENCY: str(context["currency"]),
        PLACEHOLDER_GROSS_PAY: str(context["gross_pay"]),
    }
    for worksheet in workbook.worksheets:
        for cell, text in list(_cell_texts(worksheet)):
            stripped = text.strip()
            if stripped in NUMERIC_PLACEHOLDERS:
                # A cell that is nothing but a number placeholder becomes a real
                # number, so the user's own template can keep calculating.
                cell.value = Decimal(values[stripped])
                if stripped in MONEY_PLACEHOLDERS and cell.number_format == "General":
                    # Money reads as money — but never over a format the
                    # template author chose themselves.
                    cell.number_format = MONEY_NUMBER_FORMAT
                continue
            replaced = text
            for name, value in values.items():
                replaced = replaced.replace(name, value)
            if replaced != text:
                # Substituted text is user-derived: never a live formula.
                write_text_cell(cell, replaced)


def check_template(template_path: Path) -> Workbook:
    """Load a template and verify its placeholders, writing nothing.

    Run before any document is published, so the common template mistakes are
    reported while the previous month's tally is still intact on disk.
    """
    workbook = _load_template(template_path)
    _require_placeholders(workbook, template_path)
    return workbook


def fill_template(template_path: Path, context: Mapping[str, Any]) -> Workbook:
    """Return the template workbook with every placeholder resolved."""
    workbook = check_template(template_path)
    _expand_day_rows(workbook, list(context["days"]))
    _substitute_scalars(workbook, context)
    return workbook


def write_tally_xlsx(template_path: Path, context: Mapping[str, Any], path: Path) -> Path:
    workbook = fill_template(template_path, context)
    with atomic_output_file(path, suffix=".xlsx") as tmp_path:
        workbook.save(str(tmp_path))
    return path


# --------------------------------------------------------------------------
# optional XLSX -> PDF conversion
# --------------------------------------------------------------------------


def find_converter() -> str | None:
    """Path to LibreOffice, or ``None`` when this machine has none."""
    for candidate in SOFFICE_CANDIDATES:
        found = shutil.which(candidate)
        if found:
            return found
    return None


def convert_to_pdf(source: Path, target: Path, converter: str | None = None) -> str | None:
    """Convert ``source`` to a PDF at ``target``; return a note on failure.

    Returns ``None`` on success. The conversion is a bonus, never a
    prerequisite: the built-in PDF has already been written by the time this
    runs, so a missing or unhappy LibreOffice is reported, not raised.
    """
    executable = converter or find_converter()
    if executable is None:
        return (
            "LibreOffice is not installed on this computer, so no PDF was made from the template. "
            "The built-in PDF was produced as usual; you can also open the Excel file and print it "
            "to PDF yourself."
        )
    with tempfile.TemporaryDirectory(prefix="tally-convert-") as workdir:
        # Convert inside a scratch folder: LibreOffice names its output after
        # the input stem and would otherwise land on the built-in PDF's name.
        staged = Path(workdir) / source.name
        shutil.copyfile(source, staged)
        try:
            completed = subprocess.run(
                [executable, "--headless", "--convert-to", "pdf", "--outdir", workdir, str(staged)],
                capture_output=True,
                text=True,
                timeout=SOFFICE_TIMEOUT_SECONDS,
                check=False,
            )
        except (OSError, subprocess.SubprocessError) as exc:
            return f"LibreOffice could not be started ({exc}), so no PDF was made from the template."
        produced = Path(workdir) / (staged.stem + ".pdf")
        if completed.returncode != 0 or not produced.is_file():
            detail = (completed.stderr or completed.stdout or "").strip().splitlines()
            reason = detail[-1] if detail else f"exit code {completed.returncode}"
            return f"LibreOffice could not convert the template ({reason}), so no PDF was made from it."
        with atomic_output_file(target, suffix=".pdf") as tmp_path:
            shutil.copyfile(produced, tmp_path)
    return None


# --------------------------------------------------------------------------
# orchestration
# --------------------------------------------------------------------------


def generate_tally(
    data_dir: DataDir,
    *,
    worker_id: str,
    month: str,
    template: str | None = None,
    generated_on: date | None = None,
    converter: str | None = None,
    reserve: Any = None,
) -> dict[str, Any]:
    """Produce the tally documents and report exactly what was written.

    ``reserve`` is the ``reserve_new_file`` claim used elsewhere for output
    safety; passing ``None`` (the ``--force`` path) replaces existing files.
    """
    session = load_confirmed_session(data_dir, worker_id, month)
    context = tally_context(session, generated_on=generated_on)
    template_path, template_source = resolve_template(data_dir, template)

    # Every output name is resolved through the containment check, not merely
    # joined: the filename carries a worker id, and a worker id must never be
    # able to walk out of the data folder.
    output_dir = data_dir.output_dir
    pdf_path = safe_child(output_dir, tally_pdf_filename(context["worker_id"], context["month"]))
    xlsx_path = safe_child(output_dir, tally_xlsx_filename(context["worker_id"], context["month"]))
    templated_pdf_path = safe_child(
        output_dir, templated_pdf_filename(context["worker_id"], context["month"])
    )

    # Fail on a broken template before anything is published, so a typo in a
    # placeholder cannot leave a fresh PDF beside last month's workbook.
    check_template(template_path)

    claimed: list[Path] = []
    if reserve is not None:
        # Claim every possible output up front, so an existing tally is reported
        # before anything is written rather than half-overwritten. A claim that
        # fails half way through takes the earlier claims with it, or the next
        # run would trip over empty files this run created.
        try:
            for target in (pdf_path, xlsx_path, templated_pdf_path):
                reserve(target)
                claimed.append(target)
        except BaseException:
            for target in claimed:
                target.unlink(missing_ok=True)
            raise

    notes: list[str] = []
    try:
        # Both documents are rendered before either is published: a tally whose
        # PDF and workbook disagree about the month would be worse than no
        # tally at all. The PDF is rendered first so the mandatory document is
        # never the one that gets dropped for a template problem.
        with atomic_output_files([pdf_path, xlsx_path]) as (staged_pdf, staged_xlsx):
            render_tally_pdf(context, staged_pdf)
            fill_template(template_path, context).save(str(staged_xlsx))
    except BaseException:
        for target in claimed:
            target.unlink(missing_ok=True)
        raise

    note = convert_to_pdf(xlsx_path, templated_pdf_path, converter=converter)
    if note:
        notes.append(note)
        # No templated PDF this run means there must be none on disk: an empty
        # claim, or last run's copy with last run's numbers, would both be read
        # as this month's document.
        templated_pdf_path.unlink(missing_ok=True)

    files: dict[str, str | None] = {
        "pdf": str(pdf_path),
        "xlsx": str(xlsx_path),
        "templated_pdf": None if note else str(templated_pdf_path),
    }
    return {
        "worker": {"id": context["worker_id"], "display_name": context["worker_name"]},
        "month": context["month"],
        "month_title": context["month_title"],
        "generation_date": context["generation_date"],
        "confirmed_at": context["confirmed_at"],
        "total_hours": context["total_hours"],
        "hourly_rate": context["rate"],
        "currency": context["currency"],
        "gross_pay": context["gross_pay"],
        "exact_amount": context["exact_amount"],
        "rounding_applied": context["rounding_applied"],
        "calculation": calculation_sentence(context),
        "template": {"path": str(template_path), "source": template_source},
        "files": files,
        "notes": notes,
        "days": context["days"],
    }
