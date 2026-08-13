"""Render a month layout as an editable one-page A4 XLSX sheet (spec R2).

The workbook is deliberately plain: a small header block, one row per calendar
date, grey rows for resolved off days, empty hour cells the worker fills in by
hand, and a real ``SUM`` formula in the total row so an edited sheet keeps
adding up in Excel, LibreOffice or Numbers.

Two rules are load-bearing:

* **Formula-injection safety** (plan decision 12). Everything derived from user
  input — the display name, the currency label — is written as a literal string
  cell. A name like ``=cmd()`` must land in the sheet as text, never as a
  formula. Formula cells are reserved for formulas this module generates.
* **One portrait A4 page.** Excel ignores fit-to-page unless BOTH
  ``sheet_properties.pageSetUpPr.fitToPage`` and the ``page_setup`` fields are
  set, so both are written, together with an explicit print area.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.cell.cell import TYPE_STRING, Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.worksheet import Worksheet

from .layout import MonthLayout

SHEET_TITLE = "Stundenrapport"
LABEL_WORKER = "Mitarbeiter/in"
LABEL_MONTH = "Monat"
LABEL_RATE = "Stundenlohn"
LABEL_TOTAL = "Total Stunden"
COLUMN_HEADERS = ("Datum", "Stunden", "Bemerkung")

COLUMN_WIDTHS = {"A": 18, "B": 12, "C": 34}
DATE_COLUMN, HOURS_COLUMN, NOTE_COLUMN = "A", "B", "C"
LAST_COLUMN = NOTE_COLUMN

A4_PAPER_SIZE = "9"  # openpyxl's PAPERSIZE_A4

_GREY = PatternFill(fill_type="solid", start_color="FFD9D9D9", end_color="FFD9D9D9")
_HEADER_FILL = PatternFill(fill_type="solid", start_color="FFEFEFEF", end_color="FFEFEFEF")
_THIN = Side(style="thin", color="FF999999")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_BOLD = Font(bold=True)
_TITLE_FONT = Font(bold=True, size=14)
_LEFT = Alignment(horizontal="left", vertical="center")
_CENTER = Alignment(horizontal="center", vertical="center")


def _text(cell: Cell, value: object) -> Cell:
    """Write ``value`` as a literal string cell, never as a formula.

    openpyxl decides a cell is a formula purely from a leading ``=``, so a
    worker called ``=HYPERLINK(...)`` would otherwise become live content in the
    reader's spreadsheet. Overriding ``data_type`` after assignment is what the
    file actually records (the writer emits the stored type), and the quote
    prefix makes the intent visible in Excel's own UI.
    """
    text = "" if value is None else str(value)
    if text == "":
        cell.value = None
        return cell
    cell.value = text
    cell.data_type = TYPE_STRING
    if text[:1] in ("=", "+", "-", "@"):
        cell.quotePrefix = True
    return cell


def _apply_page_setup(worksheet: Worksheet, last_row: int) -> None:
    """Portrait A4, everything on one page, print area pinned to the table."""
    worksheet.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    worksheet.page_setup.orientation = "portrait"
    worksheet.page_setup.paperSize = A4_PAPER_SIZE
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 1
    # Belt and braces: some readers look at sheet_format instead of page_setup.
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.print_area = f"{DATE_COLUMN}1:{LAST_COLUMN}{last_row}"
    worksheet.page_margins.left = 0.5
    worksheet.page_margins.right = 0.5
    worksheet.page_margins.top = 0.5
    worksheet.page_margins.bottom = 0.5


def build_workbook(
    layout: MonthLayout,
    *,
    include_rate: bool = False,
    hourly_rate: str | None = None,
    currency: str | None = None,
) -> Workbook:
    """Build the monthly sheet workbook for ``layout``.

    The hourly rate is printed only when ``include_rate`` is set (spec R2) — a
    sheet handed to an employee normally shows hours, not pay.
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Stundenrapport"

    for column, width in COLUMN_WIDTHS.items():
        worksheet.column_dimensions[column].width = width

    row = 1
    _text(worksheet.cell(row=row, column=1), SHEET_TITLE).font = _TITLE_FONT
    worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 1

    _text(worksheet.cell(row=row, column=1), LABEL_WORKER).font = _BOLD
    _text(worksheet.cell(row=row, column=2), layout.display_name)
    row += 1

    _text(worksheet.cell(row=row, column=1), LABEL_MONTH).font = _BOLD
    _text(worksheet.cell(row=row, column=2), layout.title)
    row += 1

    if include_rate:
        _text(worksheet.cell(row=row, column=1), LABEL_RATE).font = _BOLD
        rate_text = " ".join(part for part in (hourly_rate, currency) if part)
        _text(worksheet.cell(row=row, column=2), rate_text)
        row += 1

    row += 1  # one blank spacer row above the table
    header_row = row
    for index, heading in enumerate(COLUMN_HEADERS, start=1):
        cell = _text(worksheet.cell(row=header_row, column=index), heading)
        cell.font = _BOLD
        cell.fill = _HEADER_FILL
        cell.border = _BORDER
        cell.alignment = _CENTER if index == 2 else _LEFT
    row += 1

    first_day_row = row
    for day in layout.rows:
        date_cell = _text(worksheet.cell(row=row, column=1), day.label)
        date_cell.alignment = _LEFT
        hours_cell = worksheet.cell(row=row, column=2)
        hours_cell.value = None  # deliberately blank: the worker writes here
        hours_cell.alignment = _CENTER
        note_cell = worksheet.cell(row=row, column=3)
        note_cell.value = None
        for cell in (date_cell, hours_cell, note_cell):
            cell.border = _BORDER
            if not day.working:
                # One shared PatternFill instance for every grey cell.
                cell.fill = _GREY
        row += 1
    last_day_row = row - 1

    total_row = row
    total_label = _text(worksheet.cell(row=total_row, column=1), LABEL_TOTAL)
    total_label.font = _BOLD
    total_label.border = _BORDER
    total_cell = worksheet.cell(row=total_row, column=2)
    # The one generated formula in the sheet.
    total_cell.value = f"=SUM({HOURS_COLUMN}{first_day_row}:{HOURS_COLUMN}{last_day_row})"
    total_cell.font = _BOLD
    total_cell.border = _BORDER
    total_cell.alignment = _CENTER
    worksheet.cell(row=total_row, column=3).border = _BORDER

    worksheet.freeze_panes = worksheet.cell(row=first_day_row, column=1)
    _apply_page_setup(worksheet, total_row)
    return workbook


def sheet_filename(layout: MonthLayout) -> str:
    return f"{layout.worker_id}-{layout.month}.xlsx"


def write_month_sheet(
    layout: MonthLayout,
    path: Path,
    *,
    include_rate: bool = False,
    hourly_rate: str | None = None,
    currency: str | None = None,
) -> dict[str, Any]:
    """Write the monthly sheet to ``path`` and describe what was produced."""
    workbook = build_workbook(
        layout,
        include_rate=include_rate,
        hourly_rate=hourly_rate,
        currency=currency,
    )
    workbook.save(str(path))
    return {
        "path": str(path),
        "rows": len(layout.rows),
        "working_days": len(layout.working_rows),
        "off_days": len(layout.off_rows),
    }
