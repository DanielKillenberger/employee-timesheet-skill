#!/usr/bin/env python3
"""Rebuild the bundled default tally template (``assets/default-tally-template.xlsx``).

The template is a binary file, so it is generated from this readable source
rather than hand-edited: run this script whenever the layout or the wording of
the default tally workbook should change, then commit the regenerated file.

    uv run tools/make_default_tally_template.py

The placeholder contract it must satisfy is described in
``references/templates.md`` and enforced by ``tests/test_tally.py``.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.properties import PageSetupProperties

REPO_ROOT = Path(__file__).resolve().parent.parent
TARGET = REPO_ROOT / "assets" / "default-tally-template.xlsx"

HEADER_FILL = PatternFill(fill_type="solid", start_color="FFEFEFEF", end_color="FFEFEFEF")
THIN = Side(style="thin", color="FF999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BOLD = Font(bold=True)
TITLE_FONT = Font(bold=True, size=14)
NOTE_FONT = Font(size=8, color="FF555555")
LEFT = Alignment(horizontal="left", vertical="center")
CENTER = Alignment(horizontal="center", vertical="center")
WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)

DISCLAIMER = (
    "Bruttolohn = bestätigte Stunden × registrierter Stundenlohn. "
    "Abzüge, Zuschläge, Sozialversicherungen und Steuern sind darin nicht berücksichtigt."
)


def build() -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Abrechnung"
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 18

    sheet["A1"] = "Stundenabrechnung"
    sheet["A1"].font = TITLE_FONT
    sheet.merge_cells("A1:B1")

    for row, (label, placeholder) in enumerate(
        (
            ("Mitarbeiter/in", "{{worker_name}}"),
            ("Monat", "{{month_title}}"),
            ("Erstellt am", "{{generation_date}}"),
        ),
        start=3,
    ):
        sheet.cell(row=row, column=1, value=label).font = BOLD
        sheet.cell(row=row, column=2, value=placeholder).alignment = LEFT

    for column, heading in enumerate(("Datum", "Stunden"), start=1):
        cell = sheet.cell(row=7, column=column, value=heading)
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = CENTER if column == 2 else LEFT

    # The marker row: cloned once per confirmed day, styles and all.
    marker = sheet.cell(row=8, column=1, value="{{day_rows}}")
    marker.border = BORDER
    marker.alignment = LEFT
    hours = sheet.cell(row=8, column=2)
    hours.border = BORDER
    hours.alignment = CENTER

    total_label = sheet.cell(row=9, column=1, value="Total Stunden")
    total_label.font = BOLD
    total_label.border = BORDER
    total_value = sheet.cell(row=9, column=2, value="{{total_hours}}")
    total_value.font = BOLD
    total_value.border = BORDER
    total_value.alignment = CENTER

    for row, (label, placeholder) in enumerate(
        (
            ("Stundenlohn", "{{rate}}"),
            ("Währung", "{{currency}}"),
            ("Bruttolohn", "{{gross_pay}}"),
        ),
        start=11,
    ):
        sheet.cell(row=row, column=1, value=label).font = BOLD
        sheet.cell(row=row, column=2, value=placeholder).alignment = LEFT

    note = sheet.cell(row=15, column=1, value=DISCLAIMER)
    note.font = NOTE_FONT
    note.alignment = WRAP
    sheet.merge_cells("A15:B15")
    sheet.row_dimensions[15].height = 28

    sheet.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    sheet.page_setup.orientation = "portrait"
    sheet.page_setup.paperSize = "9"  # A4
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.page_margins.left = 0.5
    sheet.page_margins.right = 0.5

    return workbook


def main() -> int:
    TARGET.parent.mkdir(parents=True, exist_ok=True)
    build().save(TARGET)
    print(f"Wrote {TARGET.relative_to(REPO_ROOT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
