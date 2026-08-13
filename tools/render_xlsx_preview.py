#!/usr/bin/env python3
"""Development QA aid: render a generated XLSX onto A4 so a human can look at it.

This is **not** part of the packaged skill (only `SKILL.md`, `scripts/`,
`references/` and `assets/` ship). It exists so the eyeball checklist in
`references/reference-layout.md` can be run repeatably — and in environments
with no spreadsheet application installed — by drawing the parsed workbook onto
a real A4 page with the workbook's own column widths, row heights, fills and
fonts.

It approximates a print preview; it is not Excel's print engine. Anything it
shows clipped or spilling off the page is a genuine layout problem, and a page
that fits here still deserves one real print preview before release (AC11).

Usage::

    uv run tools/render_xlsx_preview.py <sheet.xlsx> <preview.pdf>
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

MARGIN = 36  # 0.5 inch, matching the sheet's page margins
DEFAULT_ROW_HEIGHT = 15  # points, Excel's default
GREY = (0xD9 / 255, 0xD9 / 255, 0xD9 / 255)


def column_width_points(width: float | None) -> float:
    """Excel column width (in characters) -> points, via the usual pixel rule."""
    chars = 8.43 if width is None else width
    return (chars * 7 + 5) * 0.75


def render(source: Path, target: Path) -> dict[str, float]:
    worksheet = load_workbook(source).active
    columns = [chr(ord("A") + index) for index in range(worksheet.max_column)]
    widths = [column_width_points(worksheet.column_dimensions[letter].width) for letter in columns]

    page = canvas.Canvas(str(target), pagesize=A4)
    page_width, page_height = A4
    y = page_height - MARGIN

    for row_number in range(1, worksheet.max_row + 1):
        height = worksheet.row_dimensions[row_number].height or DEFAULT_ROW_HEIGHT
        y -= height
        x = MARGIN
        for index, letter in enumerate(columns):
            cell = worksheet.cell(row=row_number, column=index + 1)
            width = widths[index]

            fill = cell.fill
            if fill is not None and fill.fill_type == "solid" and fill.fgColor.rgb not in (None, "00000000"):
                rgb = str(fill.fgColor.rgb)[-6:]
                page.setFillColorRGB(*(int(rgb[i : i + 2], 16) / 255 for i in (0, 2, 4)))
                page.rect(x, y, width, height, stroke=0, fill=1)

            if cell.border is not None and cell.border.left is not None and cell.border.left.style:
                page.setStrokeColorRGB(0.6, 0.6, 0.6)
                page.rect(x, y, width, height, stroke=1, fill=0)

            value = cell.value
            if value is not None:
                font_size = float(cell.font.size or 11)
                page.setFont("Helvetica-Bold" if cell.font.bold else "Helvetica", font_size)
                page.setFillColorRGB(0, 0, 0)
                text = str(value)
                # Draw without clipping, so an overflowing value is visible as
                # the layout bug it would be.
                if cell.alignment is not None and cell.alignment.horizontal == "center":
                    page.drawCentredString(x + width / 2, y + (height - font_size) / 2 + 2, text)
                else:
                    page.drawString(x + 3, y + (height - font_size) / 2 + 2, text)
            x += width

    page.showPage()
    page.save()
    return {
        "table_width_pt": sum(widths),
        "table_height_pt": page_height - MARGIN - y,
        "printable_width_pt": page_width - 2 * MARGIN,
        "printable_height_pt": page_height - 2 * MARGIN,
    }


def main(argv: list[str]) -> int:
    if len(argv) != 3:
        print(__doc__)
        return 2
    metrics = render(Path(argv[1]), Path(argv[2]))
    print(f"Wrote {argv[2]}")
    for key, value in metrics.items():
        print(f"  {key}: {value:.1f}")
    fits = (
        metrics["table_width_pt"] <= metrics["printable_width_pt"]
        and metrics["table_height_pt"] <= metrics["printable_height_pt"]
    )
    print(f"  fits one portrait A4 page at 100%: {fits}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
